import pandas as pd
import numpy as np
from datetime import datetime
from typing import Optional, List, Dict, Any, TYPE_CHECKING
from pathlib import Path
import os

if TYPE_CHECKING:
    from ..services.analyzer import AnalysisResult

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ReportGenerator:
    """
    报告生成器
    
    支持生成：
    - 控制台报告
    - Excel 报告
    - HTML 报告
    """
    
    def __init__(self, output_dir: str = './output'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_from_result(
        self, 
        result: 'AnalysisResult', 
        formats: List[str] = ['console', 'excel', 'html'],
        filename_prefix: str = 'trade_report'
    ) -> Dict[str, str]:
        """
        从分析结果生成报告
        
        Args:
            result: AnalysisResult 对象
            formats: 报告格式列表
            filename_prefix: 文件名前缀
            
        Returns:
            生成的文件路径字典
        """
        data = self._convert_result_to_dict(result)
        output_files = {}
        
        if 'console' in formats:
            report = self.generate_console_report(data)
            print(report)
            output_files['console'] = 'printed'
        
        if 'excel' in formats:
            filepath = self.generate_excel_report(data, f'{filename_prefix}.xlsx')
            output_files['excel'] = filepath
        
        if 'html' in formats:
            filepath = self.generate_html_report(data, f'{filename_prefix}.html')
            output_files['html'] = filepath
        
        return output_files
    
    def _convert_result_to_dict(self, result: 'AnalysisResult') -> Dict[str, Any]:
        """
        将 AnalysisResult 转换为字典格式
        """
        data = {
            'summary': result.summary,
            'profit_summary': result.profit_summary,
            'performance_metrics': result.performance_metrics,
            'positions': result.positions,
            'trade_results': result.trade_results,
            'monthly_performance': result.monthly_performance,
            'stock_performance': result.stock_performance,
        }
        return data
    
    def generate_console_report(self, data: Dict[str, Any]) -> str:
        lines = []
        lines.append("=" * 60)
        lines.append("交易分析报告")
        lines.append("=" * 60)
        
        if 'summary' in data:
            s = data['summary']
            lines.append("\n【数据概况】")
            lines.append(f"  分析模式: {s.get('mode', 'full')}")
            if s.get('stock_code'):
                lines.append(f"  股票代码: {s.get('stock_code')}")
            if s.get('start_date'):
                lines.append(f"  开始日期: {s.get('start_date')}")
            if s.get('end_date'):
                lines.append(f"  结束日期: {s.get('end_date')}")
            lines.append(f"  总记录数: {s.get('total_records', 0)}")
            if 'date_range' in s and s['date_range'][0]:
                dr = s['date_range']
                lines.append(f"  时间范围: {dr[0].strftime('%Y-%m-%d')} 至 {dr[1].strftime('%Y-%m-%d')}")
            lines.append(f"  买入次数: {s.get('buy_count', 0)}")
            lines.append(f"  卖出次数: {s.get('sell_count', 0)}")
            lines.append(f"  交易证券数: {s.get('unique_securities', 0)}")
        
        if 'profit_summary' in data and data['profit_summary']:
            p = data['profit_summary']
            lines.append("\n【账户概况】")
            lines.append(f"  账户净转入: {p.net_transfer:,.2f} 元")
            lines.append(f"  现金余额: {p.cash_balance:,.2f} 元")
            lines.append(f"  逆回购出借: {p.repo_amount:,.2f} 元")
            lines.append(f"  持仓市值: {p.stock_market_value:,.2f} 元")
            lines.append(f"  账户总资产: {p.total_assets:,.2f} 元")
            lines.append("\n【盈亏汇总】")
            lines.append(f"  账户总盈亏: {p.total_profit:,.2f} 元")
            lines.append(f"  收益率: {p.profit_rate:.2f}%")
        
        if 'performance_metrics' in data and data['performance_metrics']:
            pm = data['performance_metrics']
            lines.append("\n【绩效指标】")
            lines.append(f"  总交易次数: {pm.total_trades}")
            lines.append(f"  盈利次数: {pm.winning_trades}")
            lines.append(f"  亏损次数: {pm.losing_trades}")
            lines.append(f"  胜率: {pm.win_rate:.2f}%")
            lines.append(f"  盈亏比: {pm.profit_loss_ratio:.2f}")
            lines.append(f"  夏普比率: {pm.sharpe_ratio:.2f}")
            lines.append(f"  最大回撤: {pm.max_drawdown:.2f}%")
            lines.append(f"  平均盈利: {pm.avg_profit:,.2f} 元")
            lines.append(f"  平均亏损: {pm.avg_loss:,.2f} 元")
        
        if 'positions' in data and data['positions']:
            lines.append("\n【期末持仓】")
            for code, pos in data['positions'].items():
                market_value = pos['quantity'] * pos.get('close_price', 0)
                lines.append(f"  {code} {pos['name']}: {pos['quantity']}股, 成本价 {pos['cost_price']:.4f}, 市值 {market_value:,.2f}")
        
        lines.append("\n" + "=" * 60)
        return "\n".join(lines)
    
    def generate_excel_report(self, data: Dict[str, Any], filename: str = 'trade_report.xlsx') -> str:
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export")
        
        filepath = self.output_dir / filename
        wb = Workbook()
        
        ws = wb.active
        ws.title = "汇总"
        self._write_summary_sheet(ws, data)
        
        if 'performance_metrics' in data and data['performance_metrics']:
            ws_perf = wb.create_sheet("绩效指标")
            self._write_performance_sheet(ws_perf, data['performance_metrics'])
        
        if 'trade_results' in data and data['trade_results'] is not None and not data['trade_results'].empty:
            ws_trades = wb.create_sheet("交易明细")
            self._write_dataframe(ws_trades, data['trade_results'])
        
        if 'monthly_performance' in data and data['monthly_performance'] is not None and not data['monthly_performance'].empty:
            ws_monthly = wb.create_sheet("月度绩效")
            self._write_dataframe(ws_monthly, data['monthly_performance'])
        
        if 'stock_performance' in data and data['stock_performance'] is not None and not data['stock_performance'].empty:
            ws_stock = wb.create_sheet("股票绩效")
            self._write_dataframe(ws_stock, data['stock_performance'])
        
        wb.save(filepath)
        return str(filepath)
    
    def _write_summary_sheet(self, ws, data: Dict[str, Any]):
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=11)
        
        ws['A1'] = "交易分析报告"
        ws['A1'].font = header_font
        ws.merge_cells('A1:D1')
        
        row = 3
        
        if 'summary' in data:
            s = data['summary']
            ws[f'A{row}'] = "分析配置"
            ws[f'A{row}'].font = title_font
            row += 1
            ws[f'A{row}'] = "分析模式"
            ws[f'B{row}'] = s.get('mode', 'full')
            row += 1
            if s.get('stock_code'):
                ws[f'A{row}'] = "股票代码"
                ws[f'B{row}'] = s.get('stock_code')
                row += 1
            if s.get('start_date'):
                ws[f'A{row}'] = "开始日期"
                ws[f'B{row}'] = s.get('start_date')
                row += 1
            if s.get('end_date'):
                ws[f'A{row}'] = "结束日期"
                ws[f'B{row}'] = s.get('end_date')
                row += 1
            row += 1
        
        if 'profit_summary' in data and data['profit_summary']:
            p = data['profit_summary']
            ws[f'A{row}'] = "账户概况"
            ws[f'A{row}'].font = title_font
            row += 1
            
            items = [
                ("账户净转入", p.net_transfer),
                ("现金余额", p.cash_balance),
                ("逆回购出借", p.repo_amount),
                ("持仓市值", p.stock_market_value),
                ("账户总资产", p.total_assets),
                ("账户总盈亏", p.total_profit),
                ("收益率(%)", p.profit_rate),
            ]
            
            for label, value in items:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'B{row}'].number_format = '#,##0.00'
                row += 1
            row += 1
        
        if 'positions' in data and data['positions']:
            ws[f'A{row}'] = "期末持仓"
            ws[f'A{row}'].font = title_font
            row += 1
            ws[f'A{row}'] = "证券代码"
            ws[f'B{row}'] = "证券名称"
            ws[f'C{row}'] = "持仓数量"
            ws[f'D{row}'] = "成本价"
            ws[f'E{row}'] = "收盘价"
            ws[f'F{row}'] = "市值"
            row += 1
            
            for code, pos in data['positions'].items():
                ws[f'A{row}'] = code
                ws[f'B{row}'] = pos['name']
                ws[f'C{row}'] = pos['quantity']
                ws[f'D{row}'] = pos['cost_price']
                ws[f'E{row}'] = pos.get('close_price', 0)
                ws[f'F{row}'] = pos['quantity'] * pos.get('close_price', 0)
                row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
    
    def _write_performance_sheet(self, ws, pm):
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=11)
        
        ws['A1'] = "绩效指标"
        ws['A1'].font = header_font
        ws.merge_cells('A1:D1')
        
        row = 3
        items = [
            ("总交易次数", pm.total_trades),
            ("盈利次数", pm.winning_trades),
            ("亏损次数", pm.losing_trades),
            ("胜率(%)", pm.win_rate),
            ("盈亏比", pm.profit_loss_ratio),
            ("夏普比率", pm.sharpe_ratio),
            ("最大回撤(%)", pm.max_drawdown),
            ("平均盈利(元)", pm.avg_profit),
            ("平均亏损(元)", pm.avg_loss),
            ("总盈利(元)", pm.total_profit),
            ("总亏损(元)", pm.total_loss),
        ]
        
        for label, value in items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if isinstance(value, float):
                ws[f'B{row}'].number_format = '#,##0.00'
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
    
    def _write_dataframe(self, ws, df: pd.DataFrame):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
    
    def generate_html_report(self, data: Dict[str, Any], filename: str = 'trade_report.html') -> str:
        html_content = self._generate_html_content(data)
        filepath = self.output_dir / filename
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        return str(filepath)
    
    def _generate_html_content(self, data: Dict[str, Any]) -> str:
        html = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>交易分析报告</title>
    <style>
        body { font-family: 'Microsoft YaHei', Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        h1 { color: #333; border-bottom: 2px solid #4472C4; padding-bottom: 10px; }
        h2 { color: #4472C4; margin-top: 30px; }
        .summary-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }
        .summary-card { background: #f8f9fa; padding: 15px; border-radius: 8px; border-left: 4px solid #4472C4; }
        .summary-card h3 { margin: 0 0 10px 0; color: #666; font-size: 14px; }
        .summary-card .value { font-size: 24px; font-weight: bold; color: #333; }
        .profit { color: #28a745; }
        .loss { color: #dc3545; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background-color: #4472C4; color: white; }
        tr:hover { background-color: #f5f5f5; }
        .positive { color: #28a745; }
        .negative { color: #dc3545; }
        .footer { margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 12px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 交易分析报告</h1>
"""
        
        if 'summary' in data:
            s = data['summary']
            html += f"""
        <h2>📋 分析配置</h2>
        <div class="summary-grid">
            <div class="summary-card">
                <h3>分析模式</h3>
                <div class="value">{s.get('mode', 'full')}</div>
            </div>
            <div class="summary-card">
                <h3>总记录数</h3>
                <div class="value">{s.get('total_records', 0)}</div>
            </div>
            <div class="summary-card">
                <h3>买入次数</h3>
                <div class="value">{s.get('buy_count', 0)}</div>
            </div>
            <div class="summary-card">
                <h3>卖出次数</h3>
                <div class="value">{s.get('sell_count', 0)}</div>
            </div>
        </div>
"""
        
        if 'profit_summary' in data and data['profit_summary']:
            p = data['profit_summary']
            profit_class = 'profit' if p.total_profit >= 0 else 'loss'
            html += f"""
        <h2>💰 账户概况</h2>
        <div class="summary-grid">
            <div class="summary-card">
                <h3>账户净转入</h3>
                <div class="value">{p.net_transfer:,.2f}</div>
            </div>
            <div class="summary-card">
                <h3>现金余额</h3>
                <div class="value">{p.cash_balance:,.2f}</div>
            </div>
            <div class="summary-card">
                <h3>逆回购出借</h3>
                <div class="value">{p.repo_amount:,.2f}</div>
            </div>
            <div class="summary-card">
                <h3>持仓市值</h3>
                <div class="value">{p.stock_market_value:,.2f}</div>
            </div>
            <div class="summary-card">
                <h3>账户总资产</h3>
                <div class="value">{p.total_assets:,.2f}</div>
            </div>
            <div class="summary-card">
                <h3>账户总盈亏</h3>
                <div class="value {profit_class}">{p.total_profit:,.2f}</div>
            </div>
            <div class="summary-card">
                <h3>收益率</h3>
                <div class="value {profit_class}">{p.profit_rate:.2f}%</div>
            </div>
        </div>
"""
        
        if 'performance_metrics' in data and data['performance_metrics']:
            pm = data['performance_metrics']
            html += f"""
        <h2>� 绩效指标</h2>
        <div class="summary-grid">
            <div class="summary-card">
                <h3>总交易次数</h3>
                <div class="value">{pm.total_trades}</div>
            </div>
            <div class="summary-card">
                <h3>盈利次数</h3>
                <div class="value profit">{pm.winning_trades}</div>
            </div>
            <div class="summary-card">
                <h3>亏损次数</h3>
                <div class="value loss">{pm.losing_trades}</div>
            </div>
            <div class="summary-card">
                <h3>胜率</h3>
                <div class="value">{pm.win_rate:.2f}%</div>
            </div>
            <div class="summary-card">
                <h3>盈亏比</h3>
                <div class="value">{pm.profit_loss_ratio:.2f}</div>
            </div>
            <div class="summary-card">
                <h3>夏普比率</h3>
                <div class="value">{pm.sharpe_ratio:.2f}</div>
            </div>
            <div class="summary-card">
                <h3>最大回撤</h3>
                <div class="value loss">{pm.max_drawdown:.2f}%</div>
            </div>
        </div>
"""
        
        if 'positions' in data and data['positions']:
            html += """
        <h2>� 期末持仓</h2>
        <table>
            <tr>
                <th>证券代码</th>
                <th>证券名称</th>
                <th>持仓数量</th>
                <th>成本价</th>
                <th>收盘价</th>
                <th>市值</th>
            </tr>
"""
            for code, pos in data['positions'].items():
                market_value = pos['quantity'] * pos.get('close_price', 0)
                html += f"""
            <tr>
                <td>{code}</td>
                <td>{pos['name']}</td>
                <td>{pos['quantity']}</td>
                <td>{pos['cost_price']:.4f}</td>
                <td>{pos.get('close_price', 0):.4f}</td>
                <td>{market_value:,.2f}</td>
            </tr>
"""
            html += "        </table>\n"
        
        html += f"""
        <div class="footer">
            <p>报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
"""
        return html
